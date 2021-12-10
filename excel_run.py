import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil

class ExcelCompare:
    def __int__(self):
        return

    def excel_compare(self, m_path, m_path_old, m_path_result):
        print("Excel Compare Start")
        path = m_path + "value.xlsx"
        path_result = m_path_result + "compared_result.xlsx"

        old_df = pd.read_excel(path_result)
        df = pd.read_excel(path)

        old_df['ver'] = 'old'
        df['ver'] = 'new'

        df_con = pd.concat([old_df, df], ignore_index=True)
        changes = df_con.drop_duplicates(['종목명'], keep='last')

        df_change_new = changes[changes['ver'] == 'new'].iloc[:, :-1]
        df_change_new.sort_values(by='종목명', inplace=True)

        with pd.ExcelWriter(path_result) as writer:
            df_change_new.to_excel(writer, sheet_name='Sheet', index=False)

    def day_stock_excel_compare(self, m_path, m_path_old, m_path_result):
        print("Day_Excel Compare Start")
        shutil.copy(m_path + "Day_stock_compared_result.xlsx", m_path + "Day_stock_compared_result_old.xlsx")
        path = m_path + "/data/" + "1540.xlsx"
        path_old = m_path_old + "/data/" + "0900.xlsx"
        path_result = m_path_result + "Day_stock_compared_result.xlsx"

        old_df = pd.read_excel(path_old)
        df = pd.read_excel(path)

        old_df['ver'] = 'old'
        df['ver'] = 'new'

        df_con = pd.concat([old_df, df], ignore_index=True)
        changes = df_con.drop_duplicates(df_con.columns[:-1], keep='last')

        duplicated_list = changes[changes['종목명'].duplicated()]['종목명'].to_list()
        df_change = changes[changes['종목명'].isin(duplicated_list)]

        df_change_old = df_change[df_change['ver'] == 'old'].iloc[:, :-1]
        df_change_old.sort_values(by='종목명', inplace=True)

        df_change_new = df_change[df_change['ver'] == 'new'].iloc[:, :-1]
        df_change_new.sort_values(by='종목명', inplace=True)

        df_info_change = df_change_old.copy()

        df_info_change = df_info_change.drop(['등락률', '시가총액', '외국인비율', '거래량', 'ROE', 'PER'], axis=1)
        df_info_change['전일비차'] = ''
        df_info_change['등랼률차'] = ''
        df_info_change['시가총액차'] = ''
        df_info_change['거래량차'] = ''
        df_info_change['시총순위차'] = ''

        for i in range(len(df_change_new.index)):
            df_info_change.iloc[i, 0] = str(df_change_new.iloc[i, 0])
            df_info_change.iloc[i, 2] = str(df_change_new.iloc[i, 2])
            df_info_change.iloc[i, 3] = str(df_change_new.iloc[i, 2] - df_change_old.iloc[i, 2])
            df_info_change.iloc[i, 4] = str(df_change_new.iloc[i, 3] - df_change_old.iloc[i, 3])
            df_info_change.iloc[i, 5] = str(df_change_new.iloc[i, 4] - df_change_old.iloc[i, 4])
            df_info_change.iloc[i, 6] = str(df_change_new.iloc[i, 6] - df_change_old.iloc[i, 6])
            df_info_change.iloc[i, 7] = str(df_change_new.iloc[i, 0] - df_change_old.iloc[i, 0])

        with pd.ExcelWriter(path_result) as writer:
            df_info_change.to_excel(writer, sheet_name='Sheet', index=False)

    def data_check(self, m_path):
        path = m_path + "value.xlsx"
        df = pd.read_excel(path)
        ori_len = len(df)
        sub_len = len(df.drop_duplicates(["종목명"]))
        if ori_len == sub_len:
            return 0
        else:
            return 1

    def compact_kospi(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "코스피"
        write_ws['C1'] = "전일비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "거래량"
        write_ws['F1'] = "거래대금"
        write_ws['G1'] = "개인거래"
        write_ws['H1'] = "외국인거래"
        write_ws['I1'] = "기관거래"
        write_ws['J1'] = "프로그램거래"

        count = 2

        hourlist = [9, 10, 11, 12, 13, 14, 15, 16]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)

                time = str(hour) + str(min)

                if int(time) <= 1540:
                    path_old = m_path_old + "kospi/" + time + ".xlsx"
                    load_wb = load_workbook(path_old, data_only=True)
                    load_ws = load_wb['Sheet']

                    write_ws['A' + str(count)] = time
                    write_ws['B' + str(count)] = load_ws['A2'].value
                    write_ws['C' + str(count)] = load_ws['B2'].value
                    write_ws['D' + str(count)] = load_ws['C2'].value
                    write_ws['E' + str(count)] = load_ws['D2'].value
                    write_ws['F' + str(count)] = load_ws['E2'].value
                    write_ws['G' + str(count)] = load_ws['F2'].value
                    write_ws['H' + str(count)] = load_ws['G2'].value
                    write_ws['I' + str(count)] = load_ws['H2'].value
                    write_ws['J' + str(count)] = load_ws['I2'].value

                    count = count + 1

        write_wb.save(m_path_old + "C_kospi.xlsx")

    def compact_kodaq(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "코스닥"
        write_ws['C1'] = "전일비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "거래량"
        write_ws['F1'] = "거래대금"
        write_ws['G1'] = "개인거래"
        write_ws['H1'] = "외국인거래"
        write_ws['I1'] = "기관거래"
        write_ws['J1'] = "프로그램거래"

        count = 2

        hourlist = [9, 10, 11, 12, 13, 14, 15, 16]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)

                time = str(hour) + str(min)

                if int(time) <= 1540:
                    path_old = m_path_old + "kosdaq/" + time + ".xlsx"
                    load_wb = load_workbook(path_old, data_only=True)
                    load_ws = load_wb['Sheet']

                    write_ws['A' + str(count)] = time
                    write_ws['B' + str(count)] = load_ws['A2'].value
                    write_ws['C' + str(count)] = load_ws['B2'].value
                    write_ws['D' + str(count)] = load_ws['C2'].value
                    write_ws['E' + str(count)] = load_ws['D2'].value
                    write_ws['F' + str(count)] = load_ws['E2'].value
                    write_ws['G' + str(count)] = load_ws['F2'].value
                    write_ws['H' + str(count)] = load_ws['G2'].value
                    write_ws['I' + str(count)] = load_ws['H2'].value
                    write_ws['J' + str(count)] = load_ws['I2'].value

                    count = count + 1

        write_wb.save(m_path_old + "C_kosdaq.xlsx")

    def compact_kospi200(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "코스피200"
        write_ws['C1'] = "전일대비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "거래량"
        write_ws['F1'] = "거래대금"

        count = 2

        hourlist = [9, 10, 11, 12, 13, 14, 15, 16]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)

                time = str(hour) + str(min)

                if int(time) <= 1540:
                    path_old = m_path_old + "kospi200/" + time + ".xlsx"
                    load_wb = load_workbook(path_old, data_only=True)
                    load_ws = load_wb['Sheet']

                    write_ws['A' + str(count)] = time
                    write_ws['B' + str(count)] = load_ws['A2'].value
                    write_ws['C' + str(count)] = load_ws['B2'].value
                    write_ws['D' + str(count)] = load_ws['C2'].value
                    write_ws['E' + str(count)] = load_ws['D2'].value
                    write_ws['F' + str(count)] = load_ws['E2'].value

                    count = count + 1

        write_wb.save(m_path_old + "C_kospi200.xlsx")

    def compact_etc2(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "다우산업"
        write_ws['C1'] = "전일대비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "다우운송"
        write_ws['F1'] = "전일대비"
        write_ws['G1'] = "등락률"
        write_ws['H1'] = "나스닥종합"
        write_ws['I1'] = "전일대비"
        write_ws['J1'] = "등락률"
        write_ws['K1'] = "나스닥100"
        write_ws['L1'] = "전일대비"
        write_ws['M1'] = "등락률"
        write_ws['N1'] = "S&PO500"
        write_ws['O1'] = "전일대비"
        write_ws['P1'] = "등락률"

        count = 2

        hourlist = [23, 0, 1, 2, 3, 4, 5, 6]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)

                time = str(hour) + str(min)

                if (int(time) <= 620) or (int(time) >= 2340):
                    path_old = m_path_old + "etc2/" + time + ".xlsx"
                    load_wb = load_workbook(path_old, data_only=True)
                    load_ws = load_wb['Sheet']

                    write_ws['A' + str(count)] = time
                    write_ws['B' + str(count)] = load_ws['A2'].value
                    write_ws['C' + str(count)] = load_ws['B2'].value
                    write_ws['D' + str(count)] = load_ws['C2'].value
                    write_ws['E' + str(count)] = load_ws['D2'].value
                    write_ws['F' + str(count)] = load_ws['E2'].value
                    write_ws['G' + str(count)] = load_ws['F2'].value
                    write_ws['H' + str(count)] = load_ws['G2'].value
                    write_ws['I' + str(count)] = load_ws['H2'].value
                    write_ws['J' + str(count)] = load_ws['I2'].value
                    write_ws['K' + str(count)] = load_ws['J2'].value
                    write_ws['L' + str(count)] = load_ws['K2'].value
                    write_ws['M' + str(count)] = load_ws['L2'].value
                    write_ws['N' + str(count)] = load_ws['M2'].value
                    write_ws['O' + str(count)] = load_ws['N2'].value
                    write_ws['P' + str(count)] = load_ws['O2'].value

                    count = count + 1

        write_wb.save(m_path_old + "C_etc2.xlsx")

    def compact_future(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "선물"
        write_ws['C1'] = "전일대비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "약정수량"
        write_ws['F1'] = "약정대금"

        count = 2

        hourlist = [9, 10, 11, 12, 13, 14, 15, 16]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)

                time = str(hour) + str(min)

                if (int(time) <= 1600) and (int(time) >= 920):
                    path_old = m_path_old + "future/" + time + ".xlsx"
                    load_wb = load_workbook(path_old, data_only=True)
                    load_ws = load_wb['Sheet']

                    write_ws['A' + str(count)] = time
                    write_ws['B' + str(count)] = load_ws['A2'].value
                    write_ws['C' + str(count)] = load_ws['B2'].value
                    write_ws['D' + str(count)] = load_ws['C2'].value
                    write_ws['E' + str(count)] = load_ws['D2'].value
                    write_ws['F' + str(count)] = load_ws['E2'].value

                    count = count + 1

        write_wb.save(m_path_old + "C_future.xlsx")

    def compact_data(self, m_path_old):
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        write_ws['A1'] = "시간"
        write_ws['B1'] = "종목명"
        write_ws['C1'] = "전일비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "시가총액"
        write_ws['F1'] = "외국인비율"
        write_ws['G1'] = "거래량"
        write_ws['H1'] = "PER"
        write_ws['I1'] = "ROE"

        count = 2

        hourlist = [9, 10, 11, 12, 13, 14, 15, 16]
        minlist = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]

        df = []

        for hour in hourlist:
            for min in minlist:
                if len(str(hour)) == 1:
                    hour = "0" + str(hour)
                if len(str(min)) == 1:
                    min = "0" + str(min)
                time = str(hour) + str(min)
                if int(time) <= 1540:
                    path_old = m_path_old + "data/" + time + ".xlsx"
                    if count < 3:
                        df = pd.read_excel(path_old)
                        count = count + 1
                    if count >= 3:
                        old_df = pd.read_excel(path_old)
                        df = pd.concat([df, old_df], ignore_index=True)

        with pd.ExcelWriter(m_path_old+"C_data.xlsx") as writer:
            df.to_excel(writer, sheet_name='Sheet', index=False)
