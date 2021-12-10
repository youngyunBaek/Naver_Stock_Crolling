from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from distutils.dir_util import copy_tree
import os
import shutil
import text_run


class Crolling:
    def __int__(self):
        return

    @staticmethod
    def data_copy(time, m_path):
        print("Start copy data")
        path = m_path + "analysis/" + str(time)

        try:
            if not os.path.exists(path):
                os.makedirs(path)
        except OSError:
            print("Error: Create directory")

        shutil.copy(m_path + "C_data.xlsx", path)
        shutil.copy(m_path + "C_etc2.xlsx", path)
        shutil.copy(m_path + "C_future.xlsx", path)
        shutil.copy(m_path + "C_kosdaq.xlsx", path)
        shutil.copy(m_path + "C_kospi.xlsx", path)
        shutil.copy(m_path + "C_kospi200.xlsx", path)
        shutil.copy(m_path + "Day_compared_result.xlsx", path)
        shutil.copy(m_path + "Day_stock_compared_result.xlsx", path+time+"Day.xlsx")

    @staticmethod
    def data_save_time(time, m_path, m_path_old, check):
        if check in [0, 1]:
            path = m_path + "value.xlsx"
            path_old = m_path_old + "data/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

            path = m_path + "compared_result.xlsx"
            path_old = m_path_old + "compare/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

            path = m_path + "KOSPI.xlsx"
            path_old = m_path_old + "kospi/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

            path = m_path + "KOSPI200.xlsx"
            path_old = m_path_old + "kospi200/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

            path = m_path + "KOSDAQ.xlsx"
            path_old = m_path_old + "kosdaq/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

        if check in [1, 4]:
            path = m_path + "Future.xlsx"
            path_old = m_path_old + "future/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

        if check == 5:
            path = m_path + "ETC2.xlsx"
            path_old = m_path_old + "etc2/" + time + ".xlsx"
            load_wb = load_workbook(path, data_only=True)
            load_wb.save(path_old)

        print("Save Finish")

    @staticmethod
    def data_down(m_path_driver, m_path):
        print("Data Down Start")
        tp = text_run.Text()
        path = m_path + "value.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        i_stockCount = 1

        write_ws['A1'] = "N"
        write_ws['B1'] = "종목명"
        write_ws['C1'] = "전일비"
        write_ws['D1'] = "등락률"
        write_ws['E1'] = "시가총액"
        write_ws['F1'] = "외국인비율"
        write_ws['G1'] = "거래량"
        write_ws['H1'] = "PER"
        write_ws['I1'] = "ROE"

        for page_num in range(1, 7):
            url = "https://finance.naver.com/sise/sise_market_sum.nhn?&page=" + str(page_num)
            try:
                driver.get(url)  # for loop - page 설정
            except:
                print("39: timeout occur")
                p_error = 1

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            data = soup.select('body > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) '
                               '> div:nth-child(3) > table')
            tables = data[0]

            data = tables.select('tbody > tr')
            i_lineCount = 1
            for stock_num in range(1, 81):
                if stock_num != 1:
                    if i_lineCount < 6:
                        # print("lineCount " + str(i_lineCount))
                        # print("stock_count " + str(i_stockCount))
                        stock_name = 'tbody > tr:nth-child(' + str(stock_num) + ')'
                        data = tables.select(stock_name)  # for loop - 종목 설정

                        member = data[0]

                        data = member.select('td:nth-child(1)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        no = text
                        write_ws['A' + str(i_stockCount + 1)] = int(no)

                        data = member.select('td:nth-child(2)')
                        sub_member = data[0]
                        name = sub_member.a.string
                        write_ws['B' + str(i_stockCount + 1)] = name

                        data = member.select('td:nth-child(3)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        number = text
                        write_ws['C' + str(i_stockCount + 1)] = number

                        data = member.select('td:nth-child(5)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        percent = text
                        write_ws['D' + str(i_stockCount + 1)] = percent

                        data = member.select('td:nth-child(7)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        t_value = text
                        write_ws['E' + str(i_stockCount + 1)] = t_value

                        data = member.select('td:nth-child(9)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        fore_has = text
                        write_ws['F' + str(i_stockCount + 1)] = fore_has

                        data = member.select('td:nth-child(10)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        trade = text
                        write_ws['G' + str(i_stockCount + 1)] = trade

                        data = member.select('td:nth-child(11)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        per = text
                        write_ws['H' + str(i_stockCount + 1)] = per

                        data = member.select('td:nth-child(12)')
                        sub_member = data[0]
                        text = str(sub_member)
                        text = tp.preprocessing(text)
                        roe = text
                        write_ws['I' + str(i_stockCount + 1)] = roe

                        i_stockCount = i_stockCount + 1

                    elif i_lineCount > 7:
                        i_lineCount = 0

                    i_lineCount = i_lineCount + 1

        write_wb.save(path)

        driver.quit()

    @staticmethod
    def KOSPI_data_down(m_path_driver, m_path):
        print("KOSPI Down Start")
        tp = text_run.Text()
        path = m_path + "KOSPI.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        write_ws['A1'] = "코스피"
        write_ws['B1'] = "전일비"
        write_ws['C1'] = "등락률"
        write_ws['D1'] = "거래량"
        write_ws['E1'] = "거래대금"
        write_ws['F1'] = "개인거래"
        write_ws['G1'] = "외국인거래"
        write_ws['H1'] = "기관거래"
        write_ws['I1'] = "프로그램거래"

        url = "https://finance.naver.com/sise/sise_index.nhn?code=KOSPI"
        try:
            driver.get(url)  # for loop - page 설정
        except:
            print("39: timeout occur")
            p_error = 1

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # data = soup.select('body > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) > div:nth-child(3) > table')
        data = soup.select('body > div > div > div > div > div.box_top_sub > div.inner_sub')
        member = data[0]

        data = member.select('div.subtop_sise_detail > div > em')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        KOSPI = text
        write_ws['A2'] = KOSPI

        data = member.select('div.subtop_sise_detail > div:nth-child(1) > span')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        text = text.split(' ')
        number = text[0]
        percent = text[1]
        write_ws['B2'] = number
        write_ws['C2'] = percent

        data = member.select('div.subtop_sise_detail > table > tbody > tr:nth-child(1) > td')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        quant = text
        write_ws['D2'] = quant

        sub_member = data[1]
        text = str(sub_member)
        text = tp.preprocessing(text)
        amount = text
        write_ws['E2'] = amount

        data = member.select('dl > dd')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        person = text
        write_ws['F2'] = person

        sub_member = data[1]
        text = str(sub_member)
        text = tp.preprocessing(text)
        foreigner = text
        write_ws['G2'] = foreigner

        sub_member = data[2]
        text = str(sub_member)
        text = tp.preprocessing(text)
        company = text
        write_ws['H2'] = company

        sub_member = data[5]
        text = str(sub_member)
        text = tp.preprocessing(text)
        program = text
        write_ws['I2'] = program

        write_wb.save(path)

        driver.quit()

    @staticmethod
    def KOSDAQ_data_down(m_path_driver, m_path):
        print("KOSDAQ Down Start")
        tp = text_run.Text()
        path = m_path + "KOSDAQ.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        write_ws['A1'] = "코스닥"
        write_ws['B1'] = "전일비"
        write_ws['C1'] = "등락률"
        write_ws['D1'] = "거래량"
        write_ws['E1'] = "거래대금"
        write_ws['F1'] = "개인거래"
        write_ws['G1'] = "외국인거래"
        write_ws['H1'] = "기관거래"
        write_ws['I1'] = "프로그램거래"

        url = "https://finance.naver.com/sise/sise_index.nhn?code=KOSDAQ"
        try:
            driver.get(url)  # for loop - page 설정
        except:
            print("39: timeout occur")
            p_error = 1

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # data = soup.select('body > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) > div:nth-child(3) > table')
        data = soup.select('body > div > div > div > div > div.box_top_sub > div.inner_sub')
        member = data[0]

        data = member.select('div.subtop_sise_detail > div > em')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        KOSPI = text
        write_ws['A2'] = KOSPI

        data = member.select('div.subtop_sise_detail > div:nth-child(1) > span')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        text = text.split(' ')
        number = text[0]
        percent = text[1]
        write_ws['B2'] = number
        write_ws['C2'] = percent

        data = member.select('div.subtop_sise_detail > table > tbody > tr:nth-child(1) > td')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        quant = text
        write_ws['D2'] = quant

        sub_member = data[1]
        text = str(sub_member)
        text = tp.preprocessing(text)
        amount = text
        write_ws['E2'] = amount

        data = member.select('dl > dd')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        person = text
        write_ws['F2'] = person

        sub_member = data[1]
        text = str(sub_member)
        text = tp.preprocessing(text)
        foreigner = text
        write_ws['G2'] = foreigner

        sub_member = data[2]
        text = str(sub_member)
        text = tp.preprocessing(text)
        company = text
        write_ws['H2'] = company

        sub_member = data[5]
        text = str(sub_member)
        text = tp.preprocessing(text)
        program = text
        write_ws['I2'] = program

        write_wb.save(path)

        driver.quit()

    @staticmethod
    def KOSPI200_data_down(m_path_driver, m_path):
        print("KOSPI200 Down Start")
        tp = text_run.Text()
        path = m_path + "KOSPI200.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        write_ws['A1'] = "코스피200"
        write_ws['B1'] = "전일대비"
        write_ws['C1'] = "등락률"
        write_ws['D1'] = "거래량"
        write_ws['E1'] = "거래대금"

        url = "https://finance.naver.com/sise/sise_index.nhn?code=KPI200"
        try:
            driver.get(url)  # for loop - page 설정
        except:
            print("39: timeout occur")
            p_error = 1

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # data = soup.select('body > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) > div:nth-child(3) > table')
        data = soup.select('body > div > div > div > div > div.box_top_sub > div.inner_sub')
        member = data[0]

        data = member.select('div.subtop_sise_detail > table > tbody > tr > td')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        KOSDAQ200 = text
        write_ws['A2'] = KOSDAQ200

        sub_member = data[4]
        sub_data = sub_member.select('span')
        text = sub_data[0]
        text = str(text)
        text = tp.preprocessing(text)
        text = text.replace("</span>", "")
        value = text
        write_ws['B2'] = value

        sub_member = data[8]
        text = str(sub_member)
        text = tp.preprocessing(text)
        rate = text
        write_ws['C2'] = rate

        data = member.select('div.subtop_sise_detail > table > tbody > tr > td:nth-child(5)')
        sub_member = data[5]
        text = str(sub_member)
        text = tp.preprocessing(text)
        quant = text
        write_ws['D2'] = quant

        sub_member = data[6]
        text = str(sub_member)
        text = tp.preprocessing(text)
        amount = text
        write_ws['E2'] = amount

        write_wb.save(path)

        driver.quit()

    @staticmethod
    def Future_data_down(m_path_driver, m_path):
        print("Future Down Start")
        tp = text_run.Text()
        path = m_path + "Future.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        write_ws['A1'] = "선물"
        write_ws['B1'] = "전일대비"
        write_ws['C1'] = "등락률"
        write_ws['D1'] = "약정수량"
        write_ws['E1'] = "약정대금"

        url = "https://finance.naver.com/sise/sise_index.nhn?code=FUT"
        try:
            driver.get(url)  # for loop - page 설정
        except:
            print("39: timeout occur")
            p_error = 1

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # data = soup.select('body > div:nth-child(5) > div:nth-child(2) > div:nth-child(2) > div:nth-child(3) > table')
        data = soup.select('body > div > div > div > div > div.box_top_sub > div.inner_sub')
        member = data[0]

        data = member.select('div.subtop_sise_detail > table > tbody > tr > td:nth-child(2)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        future = text
        write_ws['A2'] = future

        sub_member = data[1]
        text = str(sub_member)
        text = tp.preprocessing(text)
        number = text
        write_ws['B2'] = number

        sub_member = data[2]
        text = str(sub_member)
        text = tp.preprocessing(text)
        percent = text
        write_ws['C2'] = percent

        sub_member = data[3]
        text = str(sub_member)
        text = tp.preprocessing(text)
        quant = text
        write_ws['D2'] = quant

        data = member.select('div.subtop_sise_detail > table > tbody > tr > td:nth-child(5)')
        sub_member = data[3]
        text = str(sub_member)
        text = tp.preprocessing(text)
        amount = text
        write_ws['E2'] = amount

        write_wb.save(path)

        driver.quit()

    @staticmethod
    def ETC2_data_down(m_path_driver, m_path):
        print("ETC2 Down Start")
        tp = text_run.Text()
        path = m_path + "ETC2.xlsx"
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.style = 'pandas'

        p_error = 0

        driver = webdriver.Chrome(m_path_driver)

        if p_error == 0:
            driver.implicitly_wait(8)

        write_ws['A1'] = "다우산업"
        write_ws['B1'] = "전일대비"
        write_ws['C1'] = "등락률"
        write_ws['D1'] = "다우운송"
        write_ws['E1'] = "전일대비"
        write_ws['F1'] = "등락률"
        write_ws['G1'] = "나스닥종합"
        write_ws['H1'] = "전일대비"
        write_ws['I1'] = "등락률"
        write_ws['J1'] = "나스닥100"
        write_ws['K1'] = "전일대비"
        write_ws['L1'] = "등락률"
        write_ws['M1'] = "S&P500"
        write_ws['N1'] = "전일대비"
        write_ws['O1'] = "등락률"

        url = "https://finance.naver.com/world/"

        try:
            driver.get(url)  # for loop - page 설정
        except:
            print("39: timeout occur")

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        data = soup.select(
            'body > div:nth-child(5) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > table > thead')
        member = data[0]

        data = member.select('tr:nth-child(2) > td:nth-child(3)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['A2'] = ETC2

        data = member.select('tr:nth-child(2) > td:nth-child(4)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['B2'] = ETC2

        data = member.select('tr:nth-child(2) > td:nth-child(5)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['C2'] = ETC2

        data = member.select('tr:nth-child(3) > td:nth-child(3)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['D2'] = ETC2

        data = member.select('tr:nth-child(3) > td:nth-child(4)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['E2'] = ETC2

        data = member.select('tr:nth-child(3) > td:nth-child(5)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['F2'] = ETC2

        data = member.select('tr:nth-child(4) > td:nth-child(3)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['G2'] = ETC2

        data = member.select('tr:nth-child(4) > td:nth-child(4)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['H2'] = ETC2

        data = member.select('tr:nth-child(4) > td:nth-child(5)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['I2'] = ETC2

        data = member.select('tr:nth-child(5) > td:nth-child(3)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['J2'] = ETC2

        data = member.select('tr:nth-child(5) > td:nth-child(4)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['K2'] = ETC2

        data = member.select('tr:nth-child(5) > td:nth-child(5)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['L2'] = ETC2

        data = member.select('tr:nth-child(6) > td:nth-child(3)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['M2'] = ETC2

        data = member.select('tr:nth-child(6) > td:nth-child(4)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['N2'] = ETC2

        data = member.select('tr:nth-child(6) > td:nth-child(5)')
        sub_member = data[0]
        text = str(sub_member)
        text = tp.preprocessing(text)
        ETC2 = text
        write_ws['O2'] = ETC2

        write_wb.save(path)

        driver.quit()